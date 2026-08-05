from __future__ import annotations

import csv
import io
from html import escape
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font as XlsxFont
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .version import APP_VERSION, DETECTORS_VERSION


# The reviewer's decision is the product's output, so every exported report
# carries the same statement of what the machine result does and does not mean.
# Efficacy is unmeasured against known-forged documents; a report that hid that
# would let a flag read as proof.
CALIBRATION_NOTICE = [
    "Parakh reports signals, not conclusions. A flagged check is a reason to look at the "
    "document, not evidence that it was forged.",
    "Detector accuracy has not been measured against known-forged documents. The "
    "true-positive rate of this system is unknown.",
    "Several checks report raw signal rather than a tuned threshold, so their flags carry "
    "no calibrated probability.",
    "A clear result is not a certificate of authenticity. It means no check in the "
    "selected set produced a signal.",
    "The recorded human decision is the verdict of record. The machine result is input to "
    "that decision, never a replacement for it.",
]

DECISION_LABELS = {
    "verified": "Verified",
    "needs_investigation": "Needs investigation",
    "inconclusive": "Inconclusive",
    "escalated": "Escalated to a second reviewer",
}

# Mirrors `analyzerLabel` in the frontend. A report that named a check
# differently from the screen the reviewer read it on would be a different
# document as far as anyone auditing the case is concerned.
_ANALYZER_LABELS = {
    "tamper_scan": "Whitener Detection",
    "photo_detection": "Document Photo",
    "qr_presence": "QR Presence",
}


def analyzer_label(analyzer_id: str) -> str:
    return _ANALYZER_LABELS.get(analyzer_id, analyzer_id.replace("_", " ").title())


_VERDICT_LABELS = {
    "review": "Needs review",
    "error": "Check errors",
    "clear": "Clear",
    "inconclusive": "Inconclusive",
    "pending": "Screening incomplete",
    "unanalyzable": "Marked unanalyzable",
}


def machine_verdict(job: dict[str, Any]) -> str:
    """Collapse a job's analyzer outcomes into one machine reading.

    Mirrors `docVerdict` in the frontend. Kept in step deliberately: an exported
    report that disagreed with the screen the reviewer looked at would be worse
    than no export at all.
    """

    if job.get("unanalyzable"):
        return "unanalyzable"
    if job.get("status") != "completed":
        return "pending"
    # `info` checks (metadata provenance) state facts and take no position, so
    # they neither clear a document nor flag it.
    outcomes = [
        str(result.get("outcome")) for result in job.get("results", {}).values()
        if str(result.get("outcome")) != "info"
    ]
    if not outcomes:
        return "inconclusive"
    if "review" in outcomes:
        return "review"
    if "error" in outcomes:
        return "error"
    if all(outcome == "clear" for outcome in outcomes):
        return "clear"
    return "inconclusive"


def verdict_label(verdict: str) -> str:
    return _VERDICT_LABELS.get(verdict, "Inconclusive")


def decision_label(decision: str | None) -> str:
    if not decision:
        return "Not yet recorded"
    return DECISION_LABELS.get(decision, decision.replace("_", " ").capitalize())


def case_report(job: dict[str, Any], batch_id: str | None = None) -> dict[str, Any]:
    """The full, machine-readable record of one screened document."""

    results = job.get("results", {})
    review = job.get("review") if isinstance(job.get("review"), dict) else None
    checks = []
    for analyzer in job.get("analyzers", []):
        result = results.get(analyzer)
        run = job.get("analyzer_runs", {}).get(analyzer, {})
        verdict = (result or {}).get("check") or {}
        checks.append({
            "analyzer_id": analyzer,
            "status": run.get("status", "queued"),
            "outcome": (result or {}).get("outcome"),
            "risk": (result or {}).get("risk"),
            "summary": (result or {}).get("summary"),
            "findings_count": (result or {}).get("findings_count", 0),
            "error": run.get("error"),
            # The check's own rule and how this document met it. Carried into the
            # export so a case file records why a tick was a tick.
            "result": verdict.get("status"),
            "criterion": verdict.get("criterion"),
            "reason": verdict.get("reason"),
            "facts": verdict.get("facts", []),
        })

    evidence = [
        {**region, "analyzer_id": analyzer}
        for analyzer, result in results.items()
        for region in result.get("regions", [])
    ]

    verdict = machine_verdict(job)
    profile = job.get("profile") if isinstance(job.get("profile"), dict) else None
    return {
        "case_id": job.get("id"),
        "batch_id": batch_id,
        "filename": job.get("filename"),
        "sha256": job.get("sha256"),
        "profile": profile,
        "screened_at": job.get("created_at"),
        "completed_at": job.get("completed_at"),
        "app_version": APP_VERSION,
        "detectors_version": DETECTORS_VERSION,
        "analysis_settings": job.get("settings", {}),
        "machine_verdict": verdict,
        "machine_verdict_label": verdict_label(verdict),
        "unanalyzable": bool(job.get("unanalyzable")),
        "unanalyzable_reason": job.get("unanalyzable_reason"),
        "checks": checks,
        "evidence_regions": evidence,
        "review": review,
        "review_label": decision_label((review or {}).get("decision")),
        "calibration_notice": CALIBRATION_NOTICE,
    }


_REPORT_CSS = """
  :root { color-scheme: light; }
  * { box-sizing: border-box; }
  body { margin: 0; padding: 40px; background: #ffffff; color: #16161c;
         font: 15px/1.55 "Inter", "Segoe UI", system-ui, sans-serif; }
  main { max-width: 920px; margin: 0 auto; }
  h1 { font-size: 24px; margin: 0 0 4px; letter-spacing: -0.01em; }
  h2 { font-size: 15px; text-transform: uppercase; letter-spacing: 0.08em;
       color: #6b6b7a; margin: 32px 0 12px; }
  .eyebrow { font-size: 12px; letter-spacing: 0.12em; text-transform: uppercase;
             color: #6b6b7a; margin: 0 0 16px; }
  .verdict { display: flex; align-items: baseline; gap: 12px; padding: 16px 18px;
             border: 1px solid #d9d9e3; border-left-width: 4px; border-radius: 6px; margin: 18px 0 0; }
  .verdict.clear { border-left-color: #1a8f86; }
  .verdict.review { border-left-color: #b8791d; }
  .verdict.error, .verdict.unanalyzable { border-left-color: #c4405f; }
  .verdict.inconclusive, .verdict.pending { border-left-color: #6b6b7a; }
  .verdict strong { font-size: 18px; }
  .verdict span { color: #5a5a68; font-size: 13px; }
  .scroll { overflow-x: auto; }
  table { width: 100%; border-collapse: collapse; font-size: 14px; }
  th, td { text-align: left; padding: 8px 10px; border-bottom: 1px solid #e6e6ee; vertical-align: top; }
  th { font-weight: 600; color: #4a4a58; }
  dl.meta { display: grid; grid-template-columns: repeat(2, minmax(0, 1fr)); gap: 10px 24px; margin: 0; }
  dl.meta div { display: flex; justify-content: space-between; gap: 16px;
                border-bottom: 1px solid #eeeef4; padding-bottom: 6px; }
  dl.meta dt { color: #6b6b7a; font-size: 13px; }
  dl.meta dd { margin: 0; font-size: 13px; text-align: right; word-break: break-all; }
  .pill { display: inline-block; padding: 2px 8px; border-radius: 999px; font-size: 12px;
          border: 1px solid #d9d9e3; }
  .pill.clear { border-color: #1a8f86; color: #14706a; }
  .pill.review { border-color: #b8791d; color: #8d5c13; }
  .pill.error { border-color: #c4405f; color: #a02d49; }
  .pill.inconclusive { border-color: #9a9aa8; color: #5a5a68; }
  td small { display: block; color: #6b6b7a; font-size: 11px; margin-top: 3px; }
  ul.facts { margin: 6px 0 0; padding-left: 16px; color: #4a4a58; font-size: 12px; }
  ul.facts b { font-weight: 600; color: #6b6b7a; }
  .auth-bar { position: relative; width: 96px; height: 8px; border-radius: 4px; background: #e6e6ee; }
  .auth-bar > span { display: block; height: 100%; border-radius: 4px; background: #9b3834; }
  .auth-bar.clear > span { background: #1a8f86; }
  .auth-bar.review > span, .auth-bar.inconclusive > span { background: #b8791d; }
  .auth-cell { display: flex; align-items: center; gap: 8px; white-space: nowrap; }
  .auth-cell b { font-size: 12px; font-variant-numeric: tabular-nums; }
  .tick-grid { text-align: center; font-size: 13px; }
  .tick-grid.pass { color: #14706a; }
  .tick-grid.fail { color: #a02d49; }
  .tick-grid.pending { color: #9a9aa8; }
  .notes { white-space: pre-wrap; border: 1px solid #e6e6ee; border-radius: 6px; padding: 12px 14px; }
  ul.notice { padding-left: 18px; color: #4a4a58; font-size: 13px; }
  ul.notice li { margin-bottom: 6px; }
  footer { margin-top: 36px; padding-top: 14px; border-top: 1px solid #e6e6ee;
           color: #6b6b7a; font-size: 12px; }
  @media (max-width: 720px) {
    body { padding: 24px 18px; }
    dl.meta { grid-template-columns: 1fr; }
    .verdict { flex-direction: column; gap: 4px; }
  }
  @media print {
    body { padding: 0; }
    h2 { break-after: avoid; }
    table { break-inside: auto; }
    .scroll { overflow-x: visible; }
  }
"""


def _document(title: str, body: str) -> str:
    return (
        "<!doctype html><html lang=\"en\"><head><meta charset=\"utf-8\">"
        f"<title>{escape(title)}</title><style>{_REPORT_CSS}</style></head>"
        f"<body><main>{body}</main></body></html>"
    )


def _meta_rows(pairs: list[tuple[str, Any]]) -> str:
    cells = "".join(
        f"<div><dt>{escape(label)}</dt><dd>{escape(str(value) if value not in (None, '') else '—')}</dd></div>"
        for label, value in pairs
    )
    return f"<dl class=\"meta\">{cells}</dl>"


# How a check result prints. `info` is not a middle grade between pass and fail
# — it is a check that reports facts and grades nothing.
_RESULT_LABELS = {
    "pass": ("Pass", "clear"),
    "fail": ("Fail", "review"),
    "inconclusive": ("Inconclusive", "inconclusive"),
    "info": ("Reported", "inconclusive"),
    "error": ("Check error", "error"),
}


def _result_label(check: dict[str, Any]) -> str:
    result = str(check.get("result") or "")
    if result in _RESULT_LABELS:
        return _RESULT_LABELS[result][0]
    return str(check.get("outcome") or check.get("status") or "—").title()


def _result_class(check: dict[str, Any]) -> str:
    return _RESULT_LABELS.get(str(check.get("result") or ""), ("", ""))[1]


def _facts_html(facts: Any) -> str:
    """The detector's own values, printed under the reason."""

    if not isinstance(facts, list) or not facts:
        return ""
    rows = "".join(
        f"<li><b>{escape(str(fact.get('label')))}:</b> {escape(str(fact.get('value')))}</li>"
        for fact in facts
        if isinstance(fact, dict) and fact.get("label")
    )
    return f"<ul class=\"facts\">{rows}</ul>" if rows else ""


def _notice_html() -> str:
    items = "".join(f"<li>{escape(line)}</li>" for line in CALIBRATION_NOTICE)
    return f"<h2>What this screening does not claim</h2><ul class=\"notice\">{items}</ul>"


def render_case_html(report: dict[str, Any]) -> str:
    """A printable single-case report. Print to PDF is the export path."""

    verdict = str(report["machine_verdict"])
    review = report.get("review") or {}

    checks = "".join(
        "<tr>"
        f"<td>{escape(analyzer_label(str(check['analyzer_id'])))}"
        f"<br><small>{escape(str(check.get('criterion') or ''))}</small></td>"
        f"<td><span class=\"pill {escape(_result_class(check))}\">"
        f"{escape(_result_label(check))}</span></td>"
        f"<td>{escape(str(check.get('reason') or check['summary'] or check['error'] or '—'))}"
        f"{_facts_html(check.get('facts'))}</td>"
        f"<td>{escape(str(check['findings_count']))}</td>"
        "</tr>"
        for check in report["checks"]
    )

    evidence = report["evidence_regions"]
    evidence_html = ""
    if evidence:
        rows = "".join(
            "<tr>"
            f"<td>{escape(str(region.get('page')))}</td>"
            f"<td>{escape(str(region.get('label')))}</td>"
            f"<td>{escape(str(region.get('severity', '')).title())}</td>"
            f"<td>{escape(str(region.get('message')))}</td>"
            "</tr>"
            for region in evidence
        )
        evidence_html = (
            "<h2>Located evidence</h2><div class=\"scroll\"><table><thead><tr><th>Page</th>"
            f"<th>Marker</th><th>Severity</th><th>Detail</th></tr></thead>"
            f"<tbody>{rows}</tbody></table></div>"
        )

    if review:
        decision_html = "<h2>Reviewer decision</h2>" + _meta_rows([
            ("Decision", report["review_label"]),
            ("Recorded", review.get("reviewed_at")),
            ("Reviewer", review.get("reviewer")),
            ("Assigned to", review.get("assigned_to")),
        ])
        if review.get("notes"):
            decision_html += f"<p class=\"notes\">{escape(str(review['notes']))}</p>"
    else:
        decision_html = (
            "<h2>Reviewer decision</h2><p class=\"notes\">No decision has been recorded for this "
            "document. This report is incomplete as a case record.</p>"
        )

    profile = report.get("profile") or {}
    profile_html = ""
    if profile:
        profile_html = "<h2>Screening test</h2>" + _meta_rows([
            ("Test", profile.get("name")),
            ("Goal", profile.get("goal")),
        ])
        if profile.get("guidance"):
            profile_html += (
                "<p class=\"notes\"><strong>Decision rule</strong><br>"
                f"{escape(str(profile['guidance']))}</p>"
            )

    unanalyzable_html = ""
    if report.get("unanalyzable"):
        unanalyzable_html = (
            "<h2>Marked unanalyzable</h2><p class=\"notes\">"
            f"{escape(str(report.get('unanalyzable_reason') or 'No reason given.'))}</p>"
        )

    body = (
        "<p class=\"eyebrow\">Parakh · Document screening case report</p>"
        f"<h1>{escape(str(report['filename']))}</h1>"
        f"<div class=\"verdict {escape(verdict)}\"><strong>{escape(str(report['machine_verdict_label']))}</strong>"
        f"<span>Machine reading · reviewer decision: {escape(str(report['review_label']))}</span></div>"
        "<h2>Case record</h2>"
        + _meta_rows([
            ("Case id", report["case_id"]),
            ("Batch id", report["batch_id"]),
            ("Screened at", report["screened_at"]),
            ("Completed at", report["completed_at"]),
            ("File digest (SHA-256)", report["sha256"]),
            ("App version", report["app_version"]),
            ("Detectors version", report["detectors_version"]),
        ])
        + profile_html
        + unanalyzable_html
        + "<h2>Checks run</h2><div class=\"scroll\"><table><thead><tr>"
          "<th>Check &amp; criterion</th><th>Result</th><th>Why</th><th>Findings</th>"
          "</tr></thead>"
          f"<tbody>{checks}</tbody></table></div>"
        + evidence_html
        + decision_html
        + _notice_html()
        + f"<footer>Generated by Parakh {escape(APP_VERSION)} · detectors "
          f"{escape(DETECTORS_VERSION)} · this report reproduces the analysis settings above.</footer>"
    )
    return _document(f"Case report — {report['filename']}", body)


def batch_csv(batch: dict[str, Any]) -> str:
    """One row per document: the spreadsheet a case file actually needs."""

    buffer = io.StringIO()
    writer = csv.writer(buffer, lineterminator="\n")
    writer.writerow([
        "document", "case_id", "sha256", "screening_test", "machine_verdict",
        "reviewer_decision", "reviewer", "assigned_to", "flagged_checks", "checks_run",
        "reviewed_at", "notes", "app_version", "detectors_version",
    ])
    for job in batch.get("jobs", []):
        review = job.get("review") if isinstance(job.get("review"), dict) else {}
        profile = job.get("profile") if isinstance(job.get("profile"), dict) else {}
        results = job.get("results", {})
        flagged = [
            analyzer for analyzer, result in results.items()
            if result.get("outcome") in ("review", "error")
        ]
        writer.writerow([
            job.get("filename"),
            job.get("id"),
            job.get("sha256") or "",
            profile.get("name") or "",
            verdict_label(machine_verdict(job)),
            decision_label(review.get("decision")),
            review.get("reviewer") or "",
            review.get("assigned_to") or "",
            "; ".join(sorted(flagged)),
            len(results),
            review.get("reviewed_at") or "",
            (review.get("notes") or "").replace("\r\n", " ").replace("\n", " "),
            APP_VERSION,
            DETECTORS_VERSION,
        ])
    return buffer.getvalue()


# One sheet per check, sourced from the same `facts` labels checks.py already
# produces (see `_facts()` calls in each `_check_*` evaluator). A fact label
# that a given result doesn't populate (e.g. a fail case with no "Pages")
# simply renders blank rather than raising — the sheet reflects whatever the
# detector actually reported, never a fabricated value.
_CHECK_SHEETS: list[tuple[str, str, list[str]]] = [
    # Metadata leads the check tabs. It grades nothing — the check is `info` by
    # design — but provenance is what a batch is read *across*: sorting a
    # hundred documents by Producer or Created clusters the ones made by one
    # hand at one sitting, which no single-document verdict can show. Excluding
    # it from the workbook because it takes no position confused "assigns no
    # blame" with "says nothing".
    ("metadata", "Metadata",
     ["Generated by", "Producer", "Creator", "PDF version", "Created", "Modified",
      "Re-saved after creation", "Pages", "Pages with a text layer", "Embedded images",
      "Annotations", "Embedded JavaScript", "Interactive form", "Title", "Author",
      "Editing signals found"]),
    ("font_analysis", "Font Analysis",
     ["Embedded fonts", "Font names", "Embedded font objects", "Referenced fonts"]),
    ("photo_detection", "Photo Module",
     ["Photos found", "Page", "Brightness", "Blur score", "Face confidence", "Quality note",
      "Search effort"]),
    ("same_phone", "Same Phone",
     ["Pages analysed", "Same-phone score", "Lowest pair score", "Findings"]),
    ("moire", "Moire",
     ["Images analysed", "File verdict", "Per-image findings"]),
    ("qr_presence", "QR Code",
     ["QR codes found", "Pages", "Decoded payloads"]),
    ("signature", "Signature",
     ["Signatures found", "Pages", "Highest confidence"]),
]

def _presence_flag(facts: dict[str, str], count_label: str) -> str:
    raw = facts.get(count_label)
    if raw is None:
        return ""
    try:
        return "Y" if int(str(raw).split()[0]) > 0 else "N"
    except ValueError:
        return ""


def _moire_flag(facts: dict[str, str]) -> str:
    verdict = facts.get("File verdict", "")
    if verdict == "Recapture":
        return "Y"
    if verdict == "Clean":
        return "N"
    return ""


def _edited_flag(facts: dict[str, str]) -> str:
    """Y when the file was written once and changed later.

    Two independent signals say so: an incremental-update structure, and a
    modification timestamp that differs from the creation timestamp. Either is
    enough. Neither is misconduct on its own — a form filled in and saved is
    exactly this — but it is the first column an auditor sorts, because a
    certificate that was edited after issue is a different object from one that
    was not. Blank when the check reported neither signal, which is not "N".
    """

    resaved = str(facts.get("Re-saved after creation", "")).strip().lower()
    created = str(facts.get("Created", "")).strip()
    modified = str(facts.get("Modified", "")).strip()
    if resaved == "yes":
        return "Y"
    if created and modified and created != modified:
        return "Y"
    if resaved == "no" or created or modified:
        return "N"
    return ""


# A derived Y/N column per sheet, read off that sheet's own facts — the
# quick-scan column, without inventing data the detector never reported (blank
# when the check has no facts to derive from, e.g. an inconclusive result).
_DERIVED_FLAGS: dict[str, tuple[str, Any]] = {
    "metadata": ("Edited after creation (Y/N)", _edited_flag),
    "photo_detection": ("Photo present (Y/N)", lambda facts: _presence_flag(facts, "Photos found")),
    "qr_presence": ("QR present (Y/N)", lambda facts: _presence_flag(facts, "QR codes found")),
    "signature": ("Signature present (Y/N)", lambda facts: _presence_flag(facts, "Signatures found")),
    "moire": ("Moire detected (Y/N)", _moire_flag),
}


# Facts are stored as display strings, which is right for the case report and
# wrong for a spreadsheet: Excel sorts text lexically, so "9%" files after
# "37%" and the borderline rows — the only ones worth reading — land in an
# arbitrary place. Every scalar a reviewer would sort or filter on is written
# back as a number here, with the display supplied by a number format.
#
# Keyed by analyzer because one label means different things on different
# sheets: "Pages" is a count on the Metadata tab and a list of page numbers on
# the Signature and QR tabs. Anything not listed stays the string the detector
# reported.
_INT, _NUMBER, _PERCENT, _SCORE = "int", "number", "percent", "score"

_NUMERIC_FACTS: dict[str, dict[str, str]] = {
    "metadata": {
        "Pages": _INT, "Pages with a text layer": _INT,
        "Embedded images": _INT, "Annotations": _INT,
    },
    "font_analysis": {"Embedded fonts": _INT, "Embedded font objects": _INT},
    "photo_detection": {
        "Photos found": _INT, "Page": _INT, "Face confidence": _PERCENT,
        "Blur score": _NUMBER, "Brightness": _NUMBER,
    },
    "same_phone": {
        "Pages analysed": _INT, "Same-phone score": _SCORE, "Lowest pair score": _NUMBER,
    },
    "moire": {"Images analysed": _INT},
    "qr_presence": {"QR codes found": _INT},
    "signature": {"Signatures found": _INT, "Highest confidence": _PERCENT},
}

# Similarity and confidence are stored as 0–1 floats so they sort numerically;
# this is how they print. One decimal place, because the difference between a
# 37% and a 34% match is exactly the distinction a reviewer is being asked to
# make.
_PERCENT_FORMAT = "0.0%"

_NUMBER_FORMATS = {
    _INT: "0",
    _NUMBER: "0.##",
    _PERCENT: _PERCENT_FORMAT,
    _SCORE: "0",
}


def _numeric_fact(value: Any, kind: str) -> Any:
    """Coerce one display string back to the number it was rendered from.

    Returns the original value untouched when it will not parse — a fact that
    reads "unknown" or carries units this does not know about is still worth
    printing, and a silently blanked cell would be worse than an unsorted one.
    """

    text = str(value).strip()
    if not text:
        return value
    try:
        if kind == _PERCENT:
            return float(text.rstrip("%")) / 100 if text.endswith("%") else float(text)
        if kind == _SCORE:
            return float(text.split("/")[0])
        number = float(text)
    except ValueError:
        return value
    return int(number) if kind == _INT and number.is_integer() else number


_RESULT_TEXT = {"pass": "Pass", "fail": "Fail", "inconclusive": "Inconclusive", "info": "Reported", "error": "Error"}


def _similarity_threshold() -> float:
    """The face-match cut-off, imported lazily.

    `photo_identity` pulls in OpenCV and the vendored photo tool at import
    time. Reporting is also used to render CSV and HTML on machines that never
    run a detector, so it must not drag that in just to print one number.
    """

    try:
        from .photo_identity import SIMILARITY_THRESHOLD

        return float(SIMILARITY_THRESHOLD)
    except Exception:
        return 0.35


def _style_header(sheet: Worksheet, headers: list[str]) -> None:
    sheet.append(headers)
    for cell in sheet[1]:
        cell.font = XlsxFont(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    for index, header in enumerate(headers, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = max(14, min(40, len(header) + 4))


def _photo_duplicate_cells(duplicate: dict[str, Any] | None) -> list[Any]:
    """The four face-matching cells: [same face as, similarity, closest, compared].

    "Not compared" and "compared, no match" are deliberately different text
    rather than both blank — a reviewer cannot act on a cell that might mean
    either. The similarity score is shown even when it falls below the
    threshold, so a near-miss is visible rather than silently discarded.

    Similarity is written as a *number*, not "19%" text, so the column sorts
    and filters as a score. Excel orders text lexically, which would file 9%
    after 19% and put the near-misses — the only rows worth reading — in an
    arbitrary place. `_PERCENT_FORMAT` supplies the display.

    "Photos compared" reports coverage as "1 of 2" whenever the detector boxed
    more photos than could be matched, because a no-match reached on half a
    document's photos is a weaker statement than one reached on all of them.
    """

    if not duplicate:
        return ["Not compared", None, "", ""]
    status = str(duplicate.get("status") or "")
    similarity = duplicate.get("similarity")
    score = float(similarity) if isinstance(similarity, (int, float)) else None
    closest = str(duplicate.get("duplicate_of_filename") or "")

    detected = duplicate.get("photos_detected")
    compared = duplicate.get("photos_compared")
    if isinstance(detected, int) and isinstance(compared, int) and detected:
        coverage = str(compared) if compared == detected else f"{compared} of {detected}"
    else:
        coverage = ""

    if status == "duplicate":
        return [closest, score, closest, coverage]
    if status == "no_match":
        return ["No match", score, closest, coverage]
    if status == "no_face":
        # The crops exist but hold no face this model will embed — normally a
        # photo_detection false positive. Saying so beats an empty cell that
        # reads as "clean".
        return ["No face to compare", None, "", coverage]
    return ["First with a photo", None, "", coverage]


def _qr_duplicate_cells(duplicate: dict[str, Any] | None) -> list[str]:
    """The two QR-matching cells: [duplicate of, matched QR value].

    Same rule as `_photo_duplicate_cells`: a document that decoded no QR code
    was never compared, and that must not read the same as a document whose
    payload was compared and found to be unique in this batch.
    """

    if not duplicate:
        return ["No QR code to compare", ""]
    if str(duplicate.get("status") or "") == "duplicate":
        return [
            str(duplicate.get("duplicate_of_filename") or ""),
            str(duplicate.get("matched_payload") or ""),
        ]
    return ["Unique in this batch", ""]


# A face score below the match threshold but above this is still shown as its
# own row. Two unrelated people in this corpus measured 0.149 and 0.194 and the
# different-person 95th percentile sits at 0.252 (see `photo_identity`), so a
# score above this is outside the population of ordinary strangers even though
# it did not clear the bar. Those are the rows an auditor most wants to eyeball.
_NEAR_MISS_FLOOR = 0.25

# How conclusive each kind of match is, strongest first. Ordering rows by this
# rather than by score means the row that settles a cluster leads it.
_MATCH_RANK = {"Identical file": 0, "Same QR payload": 1, "Same face": 2, "Face near miss": 3}


class _Clusters:
    """Union-find over documents linked by any exact-match signal.

    Three documents where A and B share a hash and B and C share a face are one
    group, not two pairs. An auditor reads a ring, so the workbook has to
    number the ring rather than leaving them to join the pairs up by eye.
    """

    def __init__(self) -> None:
        self._parent: dict[str, str] = {}

    def _find(self, key: str) -> str:
        self._parent.setdefault(key, key)
        while self._parent[key] != key:
            self._parent[key] = self._parent[self._parent[key]]
            key = self._parent[key]
        return key

    def union(self, left: str, right: str) -> None:
        left_root, right_root = self._find(left), self._find(right)
        if left_root != right_root:
            self._parent[right_root] = left_root

    def labels(self, order: list[str]) -> dict[str, str]:
        """Cluster names (C1, C2, …) numbered by first appearance in `order`."""

        names: dict[str, str] = {}
        for key in order:
            if key not in self._parent:
                continue
            root = self._find(key)
            if root not in names:
                names[root] = f"C{len(names) + 1}"
        return {key: names[self._find(key)] for key in self._parent}


def _duplicate_pairs(
    jobs: list[dict[str, Any]],
    photo_duplicates: dict[str, dict[str, Any]],
    qr_duplicates: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], dict[str, set[str]]]:
    """Every cross-document match found in this batch, as one row per pair.

    Returns the rows plus a map of job id to the kinds of signal that touched
    it, which the Summary sheet uses for its "Duplicate signals" column so a
    reviewer sees on the triage tab that a document is linked to another.

    Both members of a pair get an entry in that map even though the row is
    written once. `photo_identity` and `qr_identity` only ever compare a
    document against *earlier* ones, so an unmirrored reading would leave the
    first document of a matched pair looking untouched — which is precisely the
    document an auditor needs to find.
    """

    names = {str(job.get("id")): str(job.get("filename") or job.get("id")) for job in jobs}
    order = [str(job.get("id")) for job in jobs]
    clusters = _Clusters()
    signals: dict[str, set[str]] = {}
    pairs: list[dict[str, Any]] = []

    def link(job_id: str, other_id: str, signal: str) -> None:
        clusters.union(job_id, other_id)
        signals.setdefault(job_id, set()).add(signal)
        signals.setdefault(other_id, set()).add(signal)

    # Byte-identical resubmission. The cheapest fraud to prove and, until now,
    # the one the workbook threw away: every job already carries the digest.
    by_digest: dict[str, list[str]] = {}
    for job in jobs:
        digest = str(job.get("sha256") or "")
        if digest:
            by_digest.setdefault(digest, []).append(str(job.get("id")))
    for digest, group in by_digest.items():
        first = group[0]
        for job_id in group[1:]:
            link(job_id, first, "hash")
            pairs.append({
                "type": "Identical file",
                "job_id": job_id,
                "other_id": first,
                "score": None,
                "evidence": f"SHA-256 {digest[:16]}…",
            })

    for job_id, entry in photo_duplicates.items():
        other_id = str(entry.get("duplicate_of_job_id") or "")
        similarity = entry.get("similarity")
        score = float(similarity) if isinstance(similarity, (int, float)) else None
        if not other_id or score is None:
            continue
        if str(entry.get("status")) == "duplicate":
            link(job_id, other_id, "face")
            pairs.append({
                "type": "Same face", "job_id": job_id, "other_id": other_id,
                "score": score, "evidence": "Face embedding match",
            })
        elif score >= _NEAR_MISS_FLOOR:
            # Deliberately not clustered: this is a lead, not a link, and
            # folding it into a cluster would launder a near-miss into a claim.
            pairs.append({
                "type": "Face near miss", "job_id": job_id, "other_id": other_id,
                "score": score,
                "evidence": f"Below the {_similarity_threshold():.0%} match threshold",
            })

    for job_id, entry in qr_duplicates.items():
        other_id = str(entry.get("duplicate_of_job_id") or "")
        if str(entry.get("status")) != "duplicate" or not other_id:
            continue
        link(job_id, other_id, "QR")
        pairs.append({
            "type": "Same QR payload", "job_id": job_id, "other_id": other_id,
            "score": None,
            "evidence": str(entry.get("matched_payload") or ""),
        })

    labels = clusters.labels(order)
    for pair in pairs:
        pair["cluster"] = labels.get(pair["job_id"], "")
        pair["document"] = names.get(pair["job_id"], pair["job_id"])
        pair["other_document"] = names.get(pair["other_id"], pair["other_id"])

    # Confirmed links first, grouped by cluster; leads last, strongest first.
    # Within a cluster the strongest kind of evidence leads, not the highest
    # number: an identical digest settles the question, while a face score of
    # any size is a resemblance. Sorting on the score alone would have filed
    # the weaker row first whenever the stronger one carried no score.
    return sorted(pairs, key=lambda pair: (
        not pair["cluster"],
        int(pair["cluster"][1:]) if pair["cluster"] else 0,
        _MATCH_RANK.get(pair["type"], len(_MATCH_RANK)),
        -(pair["score"] or 0),
    )), signals


def _job_checks(job: dict[str, Any]) -> dict[str, dict[str, Any]]:
    """Per-analyzer check rows for one job, the same shape `case_report` builds."""

    results = job.get("results", {})
    rows: dict[str, dict[str, Any]] = {}
    for analyzer in job.get("analyzers", []):
        result = results.get(analyzer) or {}
        run = job.get("analyzer_runs", {}).get(analyzer, {})
        verdict = result.get("check") or {}
        rows[analyzer] = {
            "status": run.get("status", "queued"),
            "result": verdict.get("status"),
            "reason": verdict.get("reason") or result.get("summary") or run.get("error") or "",
            "facts": {
                fact.get("label"): fact.get("value")
                for fact in verdict.get("facts", [])
                if isinstance(fact, dict) and fact.get("label")
            },
        }
    return rows


def _duplicates_sheet(
    workbook: Workbook,
    pairs: list[dict[str, Any]],
    jobs: list[dict[str, Any]],
    photo_duplicates: dict[str, dict[str, Any]],
    qr_duplicates: dict[str, dict[str, Any]],
) -> None:
    """One row per matched pair of documents, grouped into clusters.

    The check tabs answer "what is wrong with this document". This tab answers
    the question a batch exists to ask — "which of these documents are the same
    document" — which no per-document row can carry, because the answer is a
    relationship rather than a property.

    The closing notes are not decoration. Face matching silently yields nothing
    when no identity model is installed, so an empty sheet has two readings —
    "nothing matched" and "nothing was compared" — and only one of them is
    good news. The sheet states which it is.
    """

    sheet = workbook.create_sheet("Duplicates")
    headers = [
        "Cluster", "Match type", "Document", "Matched document", "Score",
        "Evidence", "Case ID", "Matched case ID",
    ]
    _style_header(sheet, headers)
    score_column = get_column_letter(headers.index("Score") + 1)

    for pair in pairs:
        sheet.append([
            pair["cluster"],
            pair["type"],
            pair["document"],
            pair["other_document"],
            pair["score"],
            pair["evidence"],
            pair["job_id"],
            pair["other_id"],
        ])
        if pair["score"] is not None:
            sheet[f"{score_column}{sheet.max_row}"].number_format = _PERCENT_FORMAT

    compared = len(jobs)
    notes = []
    if not pairs:
        notes.append(
            f"No duplicate signal was found among the {compared} document"
            f"{'' if compared == 1 else 's'} in this batch."
        )
    notes.append(
        "Every document was compared against every other by SHA-256, so an identical "
        "file resubmitted under a new name appears above."
    )
    if photo_duplicates:
        notes.append(
            f"Faces were compared across the batch. \"Same face\" is claimed at or above "
            f"{_similarity_threshold():.0%}; a \"Face near miss\" row is a score above "
            f"{_NEAR_MISS_FLOOR:.0%} that did not reach it, listed as a lead and left out "
            "of the clusters. The threshold is not calibrated against known impostors."
        )
    else:
        notes.append(
            "Face matching did not run for this batch — no identity model was available, "
            "or no document carried a usable photo crop. The absence of a face row above "
            "is therefore not evidence that no two documents share a photo."
        )
    notes.append(
        "QR payloads were compared across the batch."
        if qr_duplicates else
        "No document in this batch decoded a QR code, so no payload comparison was made."
    )
    notes.append(
        "A cluster is a group of documents linked by one or more confirmed matches. "
        "A match is a reason to look at both documents together; it is not a finding."
    )
    sheet.append([])
    for note in notes:
        sheet.append([note])


# ── per-page detail ───────────────────────────────────────────────────────────
#
# Every check grades the document. That is the right unit for a verdict — a
# certificate is authentic or it is not — but it is the wrong unit for looking
# at one, because "3 fonts" over a five-page document does not say whether all
# three sit on the page carrying the seal or whether page 4 has none at all.
#
# The detectors already record where they found things; the workbook was
# collapsing that away. These helpers put it back, one row per page, reading the
# detectors' own stored payloads. Nothing here re-runs a detector, and nothing
# here infers a page number that was not reported.


def _page_number(value: Any) -> int | None:
    try:
        page = int(value)
    except (TypeError, ValueError):
        return None
    return page if page >= 1 else None


def _analyzer_raw(job: dict[str, Any], analyzer_id: str) -> dict[str, Any] | None:
    """The detector's stored payload, or None when the check never ran.

    An empty dict and None are deliberately different: the first is a check that
    ran and reported nothing page-attributable, the second is a check that was
    not part of this document's run. A sheet that printed both the same way
    would let "not looked for" read as "looked for and not found".
    """

    result = job.get("results", {}).get(analyzer_id)
    if not isinstance(result, dict):
        return None
    raw = result.get("raw")
    return raw if isinstance(raw, dict) else {}


def _counts_by_page(items: Any, key: str = "page") -> dict[int, int]:
    counts: dict[int, int] = {}
    for item in items if isinstance(items, list) else []:
        if not isinstance(item, dict):
            continue
        page = _page_number(item.get(key))
        if page:
            counts[page] = counts.get(page, 0) + 1
    return counts


def _fonts_by_page(raw: dict[str, Any]) -> dict[int, list[str]]:
    """Distinct embedded typeface names, keyed by the page they are used on."""

    from . import checks

    pages: dict[int, set[str]] = {}
    for group in checks.embedded_font_groups(raw):
        name = str(group.get("typeface") or group.get("font_name") or "").strip()
        used = group.get("pages_used")
        for value in used if isinstance(used, list) else []:
            page = _page_number(value)
            if page:
                pages.setdefault(page, set()).add(name or "(unnamed)")
    return {page: sorted(names) for page, names in pages.items()}


def _photos_by_page(raw: dict[str, Any]) -> dict[int, int]:
    """Photo counts per page, from the boxed photos the detector returned.

    Falls back to the single `page` field for older payloads that reported a
    count and a primary page but no per-photo list.
    """

    counts = _counts_by_page(raw.get("photos"))
    if counts:
        return counts
    page = _page_number(raw.get("page"))
    if page and raw.get("photo_found") is True:
        try:
            count = int(raw.get("photo_count") or 1)
        except (TypeError, ValueError):
            count = 1
        return {page: max(count, 1)}
    return {}


def _tamper_by_page(raw: dict[str, Any]) -> dict[int, tuple[float | None, int]]:
    """Per-page whitener probability and the number of regions marked on it."""

    pages: dict[int, tuple[float | None, int]] = {}
    for entry in raw.get("pages") if isinstance(raw.get("pages"), list) else []:
        if not isinstance(entry, dict):
            continue
        page = _page_number(entry.get("page"))
        if not page:
            continue
        probability = entry.get("probability")
        regions = entry.get("regions")
        pages[page] = (
            float(probability) if isinstance(probability, (int, float)) else None,
            len(regions) if isinstance(regions, list) else 0,
        )
    return pages


def _same_phone_by_page(raw: dict[str, Any]) -> dict[int, float]:
    """Each page's *worst* pairing against another page of the same document.

    The same-phone check is a comparison, so it has no score for a page on its
    own. The lowest pair score a page took part in is the honest per-page
    reading: it is how unlike the rest of the document that page looked at its
    worst, which is the page an auditor opens first.
    """

    scores: dict[int, float] = {}
    pairs = raw.get("pair_assessments")
    for pair in pairs if isinstance(pairs, list) else []:
        if not isinstance(pair, dict):
            continue
        score = pair.get("same_phone_score")
        if not isinstance(score, (int, float)):
            continue
        for key in ("page_a", "page_b"):
            page = _page_number(pair.get(key))
            if page:
                scores[page] = min(scores.get(page, float(score)), float(score))
    return scores


# Worst reading wins when several images on one page disagree: a page holding
# one recaptured image is a recaptured page whatever else sits beside it.
_MOIRE_RANK = {"RECAPTURE": 0, "INCONCLUSIVE": 1, "CLEAN": 2}


def _moire_by_page(raw: dict[str, Any]) -> dict[int, str]:
    """Per-page recapture verdicts — only for payloads that carry a page.

    The moire scanner labels each raster it pulls with the page it came from but
    does not store that label in the result, so for most documents this is empty
    and the column stays blank rather than guessing.
    """

    verdicts: dict[int, str] = {}
    images = raw.get("images")
    for image in images if isinstance(images, list) else []:
        if not isinstance(image, dict):
            continue
        page = _page_number(image.get("page"))
        verdict = str(image.get("verdict") or "").strip().upper()
        if not page or verdict not in _MOIRE_RANK:
            continue
        current = verdicts.get(page)
        if current is None or _MOIRE_RANK[verdict] < _MOIRE_RANK[current]:
            verdicts[page] = verdict
    return {page: verdict.title() for page, verdict in verdicts.items()}


def _document_page_count(job: dict[str, Any], observed: set[int]) -> int:
    """How many pages this document has, from whatever the run recorded.

    The count matters as much as the findings: without it a page that every
    check passed over silently vanishes from the sheet, and "page 3 is missing"
    reads as "page 3 was fine". Metadata states the page count outright;
    otherwise the checks that walk every page (whitener, same-phone) bound it,
    and the highest page any detector named is the floor.
    """

    candidates = set(observed)

    metadata = _analyzer_raw(job, "metadata") or {}
    generation = metadata.get("generation")
    if isinstance(generation, dict):
        pages = _page_number(generation.get("pages"))
        if pages:
            candidates.add(pages)

    for analyzer_id in ("tamper_scan", "same_phone"):
        raw = _analyzer_raw(job, analyzer_id) or {}
        pages = raw.get("pages")
        if isinstance(pages, list) and pages:
            candidates.add(len(pages))

    return max(candidates) if candidates else 0


_PER_PAGE_NOT_RUN = "—"

# The per-page columns, in the order a reviewer reads them: what is printed on
# the page (fonts), what is stuck to it (photo, signature, QR), then the two
# signals about how the page was made.
_PER_PAGE_COLUMNS = [
    "Embedded fonts", "Font names", "Photos", "Signatures", "QR codes",
    "Whitener probability", "Whitener regions", "Same-phone lowest pair score",
    "Moire",
]

_PER_PAGE_FORMATS = {
    "Embedded fonts": "0",
    "Photos": "0",
    "Signatures": "0",
    "QR codes": "0",
    "Whitener probability": _PERCENT_FORMAT,
    "Whitener regions": "0",
    "Same-phone lowest pair score": "0",
}


def _per_page_sheet(workbook: Workbook, jobs: list[dict[str, Any]]) -> None:
    """One row per page of every document: what each check found, page by page."""

    sheet = workbook.create_sheet("Per Page")
    headers = ["Document", "Case ID", "Page"] + _PER_PAGE_COLUMNS
    _style_header(sheet, headers)
    formats = {
        get_column_letter(headers.index(label) + 1): number_format
        for label, number_format in _PER_PAGE_FORMATS.items()
    }

    without_pages: list[str] = []
    for job in jobs:
        fonts_raw = _analyzer_raw(job, "font_analysis")
        photo_raw = _analyzer_raw(job, "photo_detection")
        signature_raw = _analyzer_raw(job, "signature")
        qr_raw = _analyzer_raw(job, "qr_presence")
        tamper_raw = _analyzer_raw(job, "tamper_scan")
        phone_raw = _analyzer_raw(job, "same_phone")
        moire_raw = _analyzer_raw(job, "moire")

        fonts = _fonts_by_page(fonts_raw or {})
        photos = _photos_by_page(photo_raw or {})
        signatures = _counts_by_page((signature_raw or {}).get("regions"))
        qr_codes = _counts_by_page((qr_raw or {}).get("hits"))
        tamper = _tamper_by_page(tamper_raw or {})
        phone = _same_phone_by_page(phone_raw or {})
        moire = _moire_by_page(moire_raw or {})

        observed = set().union(fonts, photos, signatures, qr_codes, tamper, phone, moire)
        page_count = _document_page_count(job, observed)
        if not page_count:
            without_pages.append(str(job.get("filename") or job.get("id")))
            continue

        for page in range(1, page_count + 1):
            names = fonts.get(page, [])
            probability, regions = tamper.get(page, (None, 0))
            row = [
                job.get("filename"),
                job.get("id"),
                page,
                len(names) if fonts_raw is not None else _PER_PAGE_NOT_RUN,
                ", ".join(names) if fonts_raw is not None else _PER_PAGE_NOT_RUN,
                photos.get(page, 0) if photo_raw is not None else _PER_PAGE_NOT_RUN,
                signatures.get(page, 0) if signature_raw is not None else _PER_PAGE_NOT_RUN,
                qr_codes.get(page, 0) if qr_raw is not None else _PER_PAGE_NOT_RUN,
                probability if tamper_raw is not None else _PER_PAGE_NOT_RUN,
                regions if tamper_raw is not None else _PER_PAGE_NOT_RUN,
                phone.get(page) if phone_raw is not None else _PER_PAGE_NOT_RUN,
                moire.get(page, "") if moire_raw is not None else _PER_PAGE_NOT_RUN,
            ]
            sheet.append(row)
            for letter, number_format in formats.items():
                cell = sheet[f"{letter}{sheet.max_row}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format

    notes = [
        "One row per page of each document. A count of 0 means that check ran and "
        f"attributed nothing to that page; \"{_PER_PAGE_NOT_RUN}\" means the check was not "
        "run for that document, so nothing was looked for.",
        "A blank cell is a check that ran but reports nothing page by page for this "
        "document — it is not a zero.",
        "The page count comes from the document's own metadata where it was recorded, "
        "and otherwise from the checks that walk every page. A document whose run "
        "recorded no page information at all is left off this sheet.",
        "Moire is blank for most documents: the recapture scanner analyses the rasters "
        "embedded in a file and does not record which page each one came from, so the "
        "absence of a verdict here is not a clean page.",
        "The same-phone column is the lowest score that page took in any pairing against "
        "another page of the same document — how unlike the rest of the document it "
        "looked at its worst, out of 100. It is a comparison, so a single-page document "
        "has none.",
        "These are the detectors' locations, not per-page verdicts. Every check in Parakh "
        "grades the whole document; this sheet says where in the document it was looking.",
    ]
    if without_pages:
        notes.insert(0, (
            f"{len(without_pages)} document(s) recorded no page information and are not "
            f"listed above: {', '.join(without_pages[:10])}"
            + ("…" if len(without_pages) > 10 else "")
        ))
    sheet.append([])
    for note in notes:
        sheet.append([note])


def batch_xlsx(
    batch: dict[str, Any],
    photo_duplicates: dict[str, dict[str, Any]] | None = None,
    qr_duplicates: dict[str, dict[str, Any]] | None = None,
) -> bytes:
    """One workbook: a summary tab, a duplicates tab, a per-page tab, and one
    tab per check type.

    The Summary tab is the triage view — one row per document, carrying the
    digest, the screening test, both verdicts, the failure count and the
    cross-document signals, so a batch can be sorted down to the rows worth
    opening. The Duplicates tab holds the relationships between documents. The
    Per Page tab breaks each document into its pages and reports what each check
    attributed to each one, which is the detail a document-level verdict has to
    throw away.

    Each check tab lists only the documents that ran that analyzer, with the
    detector's own `facts` broken into columns — the spreadsheet a reviewer
    needs to sort and filter one check's results across a whole batch, rather
    than the one-row-per-document shape `batch_csv` gives.

    `photo_duplicates` is `photo_identity.batch_photo_duplicates()`'s output,
    keyed by job id: when a document's photo matches a face already seen
    earlier in the batch, the Photo Module sheet gets a "Duplicate of" column
    pointing back to that earlier document, and a "Similarity" column with
    the match score. `qr_duplicates` is `qr_identity.batch_qr_duplicates()`'s
    output, same shape, adding the same two columns (as "Duplicate of" /
    "Matched QR value") to the QR Code sheet. Either absent (or a job with no
    entry) renders blank — it means no comparison was possible, not that no
    match was found.
    """
    photo_duplicates = photo_duplicates or {}
    qr_duplicates = qr_duplicates or {}

    workbook = Workbook()
    jobs = batch.get("jobs", [])
    columns = _analyzer_columns(batch)

    pairs, duplicate_signals = _duplicate_pairs(jobs, photo_duplicates, qr_duplicates)

    summary = workbook.active
    summary.title = "Summary"
    # Ordered for triage: what the document is, then the two verdicts, then the
    # three columns worth sorting on (score, duplicate signals, failure count),
    # then the check matrix, then the case-record columns that only matter once
    # a row has been picked out.
    leading = [
        "Document", "Case ID", "SHA-256", "Screening test", "Machine verdict",
        "Reviewer decision", "Authenticity %", "Duplicate signals", "Failed checks",
    ]
    trailing = [
        "Reviewer", "Reviewed at", "Assigned to", "Screened at", "Completed at",
        "Unanalyzable reason", "Notes",
    ]
    headers = leading + [analyzer_label(column) for column in columns] + trailing
    _style_header(summary, headers)
    for job in jobs:
        review = job.get("review") if isinstance(job.get("review"), dict) else {}
        profile = job.get("profile") if isinstance(job.get("profile"), dict) else {}
        results = job.get("results", {})
        analyzers = set(job.get("analyzers", []))
        signals = duplicate_signals.get(str(job.get("id")), set())
        row = [
            job.get("filename"),
            job.get("id"),
            job.get("sha256") or "",
            profile.get("name") or "",
            verdict_label(machine_verdict(job)),
            decision_label(review.get("decision")),
            _authenticity_percent(job),
            ", ".join(sorted(signals)),
            sum(1 for result in results.values() if result.get("outcome") in ("review", "error")),
        ]
        for column in columns:
            if column not in analyzers:
                row.append("—")
                continue
            result = results.get(column)
            if not result:
                row.append("Pending")
                continue
            glyph, _tone = _TICK_GLYPHS.get(
                str((result.get("check") or {}).get("status") or ""),
                ("✓", "pass") if result.get("outcome") == "clear" else ("✕", "fail"),
            )
            row.append(glyph)
        row += [
            review.get("reviewer") or "",
            review.get("reviewed_at") or "",
            review.get("assigned_to") or "",
            job.get("created_at") or "",
            job.get("completed_at") or "",
            job.get("unanalyzable_reason") or "",
            str(review.get("notes") or "").replace("\r\n", " ").replace("\n", " "),
        ]
        summary.append(row)

    _duplicates_sheet(workbook, pairs, jobs, photo_duplicates, qr_duplicates)
    _per_page_sheet(workbook, jobs)

    for analyzer_id, title, fact_labels in _CHECK_SHEETS:
        sheet = workbook.create_sheet(title)
        derived = _DERIVED_FLAGS.get(analyzer_id)
        headers = ["Document", "Case ID", "Result"]
        if derived:
            headers.append(derived[0])
        headers += fact_labels + ["Reason"]
        if analyzer_id == "photo_detection":
            headers += ["Same face as", "Similarity", "Closest match", "Photos compared"]
        elif analyzer_id == "qr_presence":
            headers += ["Duplicate of", "Matched QR value"]
        _style_header(sheet, headers)

        numeric = _NUMERIC_FACTS.get(analyzer_id, {})
        formats = {
            get_column_letter(headers.index(label) + 1): _NUMBER_FORMATS[kind]
            for label, kind in numeric.items() if label in headers
        }
        if analyzer_id == "photo_detection":
            formats[get_column_letter(headers.index("Similarity") + 1)] = _PERCENT_FORMAT

        for job in jobs:
            checks = _job_checks(job)
            check = checks.get(analyzer_id)
            if check is None:
                continue
            facts = check["facts"]
            row = [job.get("filename"), job.get("id"), _RESULT_TEXT.get(str(check["result"]), check["result"] or "—")]
            if derived:
                row.append(derived[1](facts))
            row += [
                _numeric_fact(facts[label], numeric[label])
                if label in numeric and label in facts else facts.get(label, "")
                for label in fact_labels
            ]
            row.append(check["reason"])
            if analyzer_id == "photo_detection":
                row += _photo_duplicate_cells(photo_duplicates.get(str(job.get("id"))))
            elif analyzer_id == "qr_presence":
                row += _qr_duplicate_cells(qr_duplicates.get(str(job.get("id"))))
            sheet.append(row)
            for letter, number_format in formats.items():
                cell = sheet[f"{letter}{sheet.max_row}"]
                if isinstance(cell.value, (int, float)):
                    cell.number_format = number_format

        if analyzer_id == "photo_detection":
            # The score is meaningless without the line it is judged against,
            # and the workbook travels away from the app that produced it.
            sheet.append([])
            sheet.append([
                f"Similarity is a face-match score between 0% and 100%. "
                f"\"Same face as\" is filled in at or above {_similarity_threshold():.0%}; "
                f"below it the closest document is still shown so a near-miss can be "
                f"judged. The threshold is not calibrated against known impostors."
            ])

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _batch_profile(batch: dict[str, Any]) -> str | None:
    """The screening test a batch ran under, when every document shares one."""

    names = {
        (job.get("profile") or {}).get("name")
        for job in batch.get("jobs", [])
        if isinstance(job.get("profile"), dict)
    }
    names.discard(None)
    if len(names) == 1:
        return next(iter(names))
    if len(names) > 1:
        return "Mixed"
    return None


def _authenticity_percent(job: dict[str, Any]) -> int | None:
    """Share of completed checks that came back clear. A plain count, not a
    probability — mirrors `authenticityPercent` in the frontend so the report
    never disagrees with the screen the reviewer read it on."""

    graded = [
        result for result in job.get("results", {}).values()
        if result.get("outcome") != "info"
    ]
    if not graded:
        return None
    clear = sum(1 for result in graded if result.get("outcome") == "clear")
    return round((clear / len(graded)) * 100)


def _analyzer_columns(batch: dict[str, Any]) -> list[str]:
    columns: list[str] = []
    for job in batch.get("jobs", []):
        for analyzer in job.get("analyzers", []):
            if analyzer not in columns:
                columns.append(analyzer)
    return columns


# Five distinct marks, because collapsing "could not run" into a cross is the
# thing that makes a check matrix unreadable.
_TICK_GLYPHS = {
    "pass": ("✓", "pass"),
    "fail": ("✕", "fail"),
    "inconclusive": ("?", "pending"),
    "info": ("i", "pending"),
    "error": ("E", "fail"),
}


def _authenticity_chart_html(batch: dict[str, Any], columns: list[str]) -> str:
    """The tick/cross check matrix plus an authenticity bar per document —
    the same signal the batch matrix screen shows, reproduced for the printed
    report so a case file carries the chart, not just the verdict pill."""

    if not columns:
        return ""

    header_cells = "".join(f"<th>{escape(analyzer_label(column))}</th>" for column in columns)
    rows = ""
    for job in batch.get("jobs", []):
        results = job.get("results", {})
        analyzers = set(job.get("analyzers", []))
        cells = ""
        for column in columns:
            if column not in analyzers:
                cells += "<td class=\"tick-grid\">—</td>"
                continue
            result = results.get(column)
            if not result:
                cells += "<td class=\"tick-grid pending\">·</td>"
                continue
            glyph, tone = _TICK_GLYPHS.get(
                str((result.get("check") or {}).get("status") or ""),
                ("✓", "pass") if result.get("outcome") == "clear" else ("✕", "fail"),
            )
            cells += f"<td class=\"tick-grid {tone}\">{glyph}</td>"
        score = _authenticity_percent(job)
        verdict = machine_verdict(job)
        if score is None:
            auth_cell = "—"
        else:
            auth_cell = (
                f"<div class=\"auth-cell\"><div class=\"auth-bar {escape(verdict)}\">"
                f"<span style=\"width:{score}%\"></span></div><b>{score}%</b></div>"
            )
        rows += (
            "<tr>"
            f"<td>{escape(str(job.get('filename')))}</td>"
            f"<td>{auth_cell}</td>"
            f"{cells}"
            "</tr>"
        )

    return (
        "<h2>Authenticity &amp; check chart</h2>"
        "<div class=\"scroll\"><table><thead><tr><th>Document</th><th>Authenticity</th>"
        f"{header_cells}</tr></thead><tbody>{rows}</tbody></table></div>"
        "<p style=\"color:#6b6b7a;font-size:11px;margin:8px 0 0;\">✓ criterion met · "
        "✕ criterion not met · ? could not be determined · i reported, not graded · "
        "E check error · · pending · — not requested for that document</p>"
    )


def render_batch_html(batch: dict[str, Any]) -> str:
    """A printable batch summary: what got screened, what a human decided."""

    rows = ""
    counts = {"clear": 0, "review": 0, "error": 0, "inconclusive": 0, "pending": 0, "unanalyzable": 0}
    undecided = 0
    for job in batch.get("jobs", []):
        verdict = machine_verdict(job)
        counts[verdict] = counts.get(verdict, 0) + 1
        review = job.get("review") if isinstance(job.get("review"), dict) else {}
        if not review:
            undecided += 1
        rows += (
            "<tr>"
            f"<td>{escape(str(job.get('filename')))}</td>"
            f"<td><span class=\"pill {escape(verdict)}\">{escape(verdict_label(verdict))}</span></td>"
            f"<td>{escape(decision_label(review.get('decision')))}</td>"
            f"<td>{escape(str(review.get('assigned_to') or '—'))}</td>"
            f"<td>{escape(str(job.get('id'))[:12])}</td>"
            "</tr>"
        )

    chart_html = _authenticity_chart_html(batch, _analyzer_columns(batch))

    body = (
        "<p class=\"eyebrow\">Parakh · Batch screening report</p>"
        f"<h1>Batch {escape(str(batch.get('id'))[:8])}</h1>"
        f"<div class=\"verdict {'review' if counts['review'] else 'clear'}\">"
        f"<strong>{len(batch.get('jobs', []))} documents screened</strong>"
        f"<span>{counts['review']} need review · {counts['error']} check errors · "
        f"{undecided} without a recorded decision</span></div>"
        "<h2>Batch record</h2>"
        + _meta_rows([
            ("Batch id", batch.get("id")),
            ("Created at", batch.get("created_at")),
            ("Status", str(batch.get("status", "")).title()),
            ("Documents", batch.get("document_count")),
            ("Screening test", _batch_profile(batch)),
            ("App version", APP_VERSION),
            ("Detectors version", DETECTORS_VERSION),
        ])
        + "<h2>Documents</h2><div class=\"scroll\"><table><thead><tr><th>Document</th>"
          "<th>Machine verdict</th><th>Reviewer decision</th><th>Assigned to</th>"
          f"<th>Case id</th></tr></thead><tbody>{rows}</tbody></table></div>"
        + chart_html
        + _notice_html()
        + f"<footer>Generated by Parakh {escape(APP_VERSION)} · detectors {escape(DETECTORS_VERSION)}.</footer>"
    )
    return _document(f"Batch report — {batch.get('id')}", body)

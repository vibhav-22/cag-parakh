from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import load_workbook

from backend.models import normalize_analyzer_result
from backend.reporting import batch_xlsx

_RAW = {
    "font_analysis": {"typefaces": [{"typeface": "Arial"}]},
    "photo_detection": {
        "photo_found": True,
        "photo_count": 1,
        "passed": True,
        "brightness": 128,
        "blur": 42,
        "face_confidence": 0.9,
        "page": 1,
    },
    "same_phone": {
        "summary": {"overall_verdict": "likely_same_phone_or_workflow", "same_phone_score": 87},
        "pages": [{"page": 1}, {"page": 2}],
    },
    "moire": {"file_verdict": "CLEAN", "images": [{"page": 1}]},
    "qr_presence": {"qr_count": 1, "hits": [{"page": 1, "data": "https://example.com"}]},
    "signature": {"count": 1, "regions": [{"page": 1, "confidence": 0.8}]},
}


def _job() -> dict:
    results = {
        analyzer_id: normalize_analyzer_result(analyzer_id, raw).model_dump(mode="json")
        for analyzer_id, raw in _RAW.items()
    }
    analyzer_runs = {
        analyzer_id: {"status": "completed", "error": None}
        for analyzer_id in _RAW
    }
    return {
        "id": "job-1",
        "filename": "sample.pdf",
        "analyzers": list(_RAW),
        "analyzer_runs": analyzer_runs,
        "results": results,
        "review": {"decision": "verified"},
    }


class BatchXlsxTests(unittest.TestCase):
    def test_workbook_has_a_summary_and_one_tab_per_check(self) -> None:
        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [_job()]})))

        self.assertEqual(
            workbook.sheetnames,
            ["Summary", "Duplicates", "Per Page", "Metadata", "Font Analysis",
             "Photo Module", "Same Phone", "Moire", "QR Code", "Signature"],
        )

    def test_font_analysis_sheet_lists_the_embedded_font_names(self) -> None:
        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [_job()]})))
        sheet = workbook["Font Analysis"]

        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["Document"], "sample.pdf")
        self.assertEqual(row["Result"], "Pass")
        self.assertEqual(row["Font names"], "Arial")

    def test_photo_module_sheet_derives_a_presence_flag(self) -> None:
        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [_job()]})))
        sheet = workbook["Photo Module"]

        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["Photo present (Y/N)"], "Y")
        # Written as a number so the column sorts as a measurement, not as text.
        self.assertEqual(row["Brightness"], 128)

    def test_scores_are_written_as_sortable_numbers(self) -> None:
        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [_job()]})))

        photo = workbook["Photo Module"]
        photo_headers = [cell.value for cell in photo[1]]
        photo_cells = {header: cell for header, cell in zip(photo_headers, photo[2])}
        # "90%" text would sort after "9%"; the float sorts where a reader expects.
        self.assertAlmostEqual(photo_cells["Face confidence"].value, 0.9)
        self.assertEqual(photo_cells["Face confidence"].number_format, "0.0%")
        self.assertEqual(photo_cells["Photos found"].value, 1)

        phone = workbook["Same Phone"]
        phone_headers = [cell.value for cell in phone[1]]
        phone_row = {header: cell.value for header, cell in zip(phone_headers, phone[2])}
        # Stored by the detector as "87/100"; the denominator is not data.
        self.assertEqual(phone_row["Same-phone score"], 87)
        self.assertEqual(phone_row["Pages analysed"], 2)

    def test_a_fact_that_will_not_parse_is_left_as_reported(self) -> None:
        """A blanked cell would be worse than an unsorted one."""

        job = _job()
        job["results"]["signature"]["check"]["facts"] = [
            {"label": "Signatures found", "value": "unknown"},
        ]

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Signature"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["Signatures found"], "unknown")

    def test_photo_module_says_not_compared_when_matching_did_not_run(self) -> None:
        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [_job()]})))
        sheet = workbook["Photo Module"]

        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        # "Not compared" must never render the same as "compared, no match".
        self.assertEqual(row["Same face as"], "Not compared")
        self.assertIsNone(row["Similarity"])

    def test_photo_module_sheet_links_a_matched_duplicate(self) -> None:
        first = _job()
        first["id"], first["filename"] = "job-1", "first.pdf"
        second = _job()
        second["id"], second["filename"] = "job-2", "second.pdf"
        duplicates = {
            "job-1": {"status": "first", "similarity": None, "duplicate_of_filename": None},
            "job-2": {
                "status": "duplicate",
                "duplicate_of_job_id": "job-1",
                "duplicate_of_filename": "first.pdf",
                "similarity": 0.87,
            },
        }

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [first, second]}, duplicates)))
        sheet = workbook["Photo Module"]
        headers = [cell.value for cell in sheet[1]]
        first_row = {header: cell.value for header, cell in zip(headers, sheet[2])}
        second_row = {header: cell.value for header, cell in zip(headers, sheet[3])}

        self.assertEqual(first_row["Same face as"], "First with a photo")
        self.assertEqual(second_row["Same face as"], "first.pdf")
        # Written as a number, not "87%" text, so the column sorts as a score.
        self.assertAlmostEqual(second_row["Similarity"], 0.87)

    def test_photo_module_shows_a_below_threshold_near_miss(self) -> None:
        """A 38% near-miss must be visible, not silently dropped."""

        first = _job()
        first["id"], first["filename"] = "job-1", "first.pdf"
        second = _job()
        second["id"], second["filename"] = "job-2", "second.pdf"
        duplicates = {
            "job-2": {
                "status": "no_match",
                "duplicate_of_job_id": "job-1",
                "duplicate_of_filename": "first.pdf",
                "similarity": 0.3779,
            },
        }

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [first, second]}, duplicates)))
        sheet = workbook["Photo Module"]
        headers = [cell.value for cell in sheet[1]]
        cells = {header: cell for header, cell in zip(headers, sheet[3])}
        second_row = {header: cell.value for header, cell in cells.items()}

        self.assertEqual(second_row["Same face as"], "No match")
        self.assertAlmostEqual(second_row["Similarity"], 0.3779)
        self.assertEqual(cells["Similarity"].number_format, "0.0%")
        self.assertEqual(second_row["Closest match"], "first.pdf")

    def test_photo_module_reports_partial_photo_coverage(self) -> None:
        """A no-match reached on 1 of 2 photos is a weaker claim than one
        reached on both, and the reviewer has to be able to see the difference."""

        first = _job()
        first["id"], first["filename"] = "job-1", "first.pdf"
        second = _job()
        second["id"], second["filename"] = "job-2", "second.pdf"
        duplicates = {
            "job-1": {
                "status": "first", "similarity": None, "duplicate_of_filename": None,
                "photos_detected": 2, "photos_compared": 2,
            },
            "job-2": {
                "status": "no_match",
                "duplicate_of_job_id": "job-1",
                "duplicate_of_filename": "first.pdf",
                "similarity": 0.19,
                "photos_detected": 2,
                "photos_compared": 1,
            },
        }

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [first, second]}, duplicates)))
        sheet = workbook["Photo Module"]
        headers = [cell.value for cell in sheet[1]]
        first_row = {header: cell.value for header, cell in zip(headers, sheet[2])}
        second_row = {header: cell.value for header, cell in zip(headers, sheet[3])}

        self.assertEqual(first_row["Photos compared"], "2")
        self.assertEqual(second_row["Photos compared"], "1 of 2")

    def test_photo_module_distinguishes_a_crop_with_no_usable_face(self) -> None:
        """Photos were found but none held a face the model would embed —
        different from "no photo" and from "compared, no match"."""

        job = _job()
        job["id"], job["filename"] = "job-1", "only.pdf"
        duplicates = {
            "job-1": {
                "status": "no_face", "similarity": None, "duplicate_of_filename": None,
                "photos_detected": 1, "photos_compared": 0,
            },
        }

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]}, duplicates)))
        sheet = workbook["Photo Module"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["Same face as"], "No face to compare")
        self.assertEqual(row["Photos compared"], "0 of 1")
        self.assertIsNone(row["Similarity"])

    def test_qr_sheet_links_a_matched_duplicate(self) -> None:
        first = _job()
        first["id"], first["filename"] = "job-1", "first.pdf"
        second = _job()
        second["id"], second["filename"] = "job-2", "second.pdf"
        qr_duplicates = {
            "job-2": {
                "status": "duplicate",
                "duplicate_of_job_id": "job-1",
                "duplicate_of_filename": "first.pdf",
                "matched_payload": "https://example.com",
            },
        }

        workbook = load_workbook(
            BytesIO(batch_xlsx({"jobs": [first, second]}, qr_duplicates=qr_duplicates))
        )
        sheet = workbook["QR Code"]
        headers = [cell.value for cell in sheet[1]]
        first_row = {header: cell.value for header, cell in zip(headers, sheet[2])}
        second_row = {header: cell.value for header, cell in zip(headers, sheet[3])}

        self.assertIn("Duplicate of", headers)
        self.assertIn("Matched QR value", headers)
        self.assertEqual(second_row["Duplicate of"], "first.pdf")
        self.assertEqual(second_row["Matched QR value"], "https://example.com")
        # job-1 has no entry: it decoded no QR, so it was never compared. That
        # must not read the same as a payload that was compared and was unique.
        self.assertEqual(first_row["Duplicate of"], "No QR code to compare")

    def test_qr_sheet_marks_a_compared_but_unique_payload(self) -> None:
        job = _job()
        job["id"], job["filename"] = "job-1", "only.pdf"
        duplicates = {"job-1": {"status": "new", "duplicate_of_filename": None}}

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]}, qr_duplicates=duplicates)))
        sheet = workbook["QR Code"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["Duplicate of"], "Unique in this batch")

    def test_summary_carries_the_case_record_columns(self) -> None:
        job = _job()
        job["sha256"] = "a" * 64
        job["profile"] = {"name": "Scholarship intake"}
        job["created_at"] = "2026-07-29T10:00:00Z"
        job["review"] = {
            "decision": "verified", "reviewer": "R. Iyer",
            "reviewed_at": "2026-07-29T11:00:00Z", "notes": "Checked\nagainst register.",
        }

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Summary"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["SHA-256"], "a" * 64)
        self.assertEqual(row["Screening test"], "Scholarship intake")
        self.assertEqual(row["Reviewer"], "R. Iyer")
        self.assertEqual(row["Screened at"], "2026-07-29T10:00:00Z")
        # Newlines would break the row apart when the sheet is re-exported.
        self.assertEqual(row["Notes"], "Checked against register.")
        self.assertEqual(row["Failed checks"], 0)

    def test_summary_counts_failed_checks_for_triage(self) -> None:
        job = _job()
        job["results"]["moire"]["outcome"] = "review"
        job["results"]["signature"]["outcome"] = "error"

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Summary"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["Failed checks"], 2)

    def test_metadata_sheet_reports_provenance_and_an_edit_flag(self) -> None:
        job = _job()
        job["analyzers"].append("metadata")
        job["analyzer_runs"]["metadata"] = {"status": "completed", "error": None}
        job["results"]["metadata"] = normalize_analyzer_result("metadata", {
            "generation": {"kind": "image_editor", "producer": "Adobe Photoshop",
                           "pages": 2, "resaved_after_creation": True},
            "document_metadata": {"created_at": "2026-01-02", "modified_at": "2026-05-09"},
        }).model_dump(mode="json")

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Metadata"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["Producer"], "Adobe Photoshop")
        self.assertEqual(row["Created"], "2026-01-02")
        self.assertEqual(row["Modified"], "2026-05-09")
        self.assertEqual(row["Edited after creation (Y/N)"], "Y")
        self.assertEqual(row["Pages"], 2)
        # Info-only: the check grades nothing, and the sheet must not imply it did.
        self.assertEqual(row["Result"], "Reported")

    def test_metadata_edit_flag_is_blank_when_nothing_was_recorded(self) -> None:
        """A file whose provenance could not be read is not a file that was
        never edited, and the column must not claim it was."""

        job = _job()
        job["analyzers"].append("metadata")
        job["analyzer_runs"]["metadata"] = {"status": "completed", "error": None}
        job["results"]["metadata"] = normalize_analyzer_result(
            "metadata", {"note": "This run did not record the document's metadata."}
        ).model_dump(mode="json")

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Metadata"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertIn(row["Edited after creation (Y/N)"], (None, ""))

    def test_duplicates_sheet_pairs_byte_identical_files(self) -> None:
        first, second = _job(), _job()
        first["id"], first["filename"], first["sha256"] = "job-1", "first.pdf", "d" * 64
        second["id"], second["filename"], second["sha256"] = "job-2", "resubmitted.pdf", "d" * 64

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [first, second]})))
        sheet = workbook["Duplicates"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["Match type"], "Identical file")
        self.assertEqual(row["Document"], "resubmitted.pdf")
        self.assertEqual(row["Matched document"], "first.pdf")
        self.assertEqual(row["Cluster"], "C1")
        self.assertIn("dddddddddddddddd", row["Evidence"])

    def test_duplicates_sheet_clusters_documents_linked_through_a_third(self) -> None:
        """A and B share a hash, B and C share a face: that is one ring of
        three, not two unrelated pairs."""

        jobs = []
        for index, name in enumerate(["a.pdf", "b.pdf", "c.pdf"], start=1):
            job = _job()
            job["id"], job["filename"] = f"job-{index}", name
            jobs.append(job)
        jobs[0]["sha256"] = jobs[1]["sha256"] = "e" * 64
        photo_duplicates = {
            "job-3": {"status": "duplicate", "duplicate_of_job_id": "job-2",
                      "duplicate_of_filename": "b.pdf", "similarity": 0.81},
        }

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": jobs}, photo_duplicates)))
        sheet = workbook["Duplicates"]
        headers = [cell.value for cell in sheet[1]]
        rows = [
            {header: cell.value for header, cell in zip(headers, row)}
            for row in sheet.iter_rows(min_row=2, max_row=3)
        ]

        self.assertEqual({row["Cluster"] for row in rows}, {"C1"})
        self.assertEqual({row["Match type"] for row in rows}, {"Identical file", "Same face"})

    def test_duplicates_sheet_lists_a_near_miss_outside_any_cluster(self) -> None:
        first, second = _job(), _job()
        first["id"], first["filename"] = "job-1", "first.pdf"
        second["id"], second["filename"] = "job-2", "second.pdf"
        photo_duplicates = {
            "job-2": {"status": "no_match", "duplicate_of_job_id": "job-1",
                      "duplicate_of_filename": "first.pdf", "similarity": 0.31},
        }

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [first, second]}, photo_duplicates)))
        sheet = workbook["Duplicates"]
        headers = [cell.value for cell in sheet[1]]
        cells = {header: cell for header, cell in zip(headers, sheet[2])}

        self.assertEqual(cells["Match type"].value, "Face near miss")
        self.assertAlmostEqual(cells["Score"].value, 0.31)
        self.assertEqual(cells["Score"].number_format, "0.0%")
        # A lead is not a link: clustering it would launder it into a claim.
        self.assertIn(cells["Cluster"].value, (None, ""))

    def test_duplicates_sheet_ignores_a_score_below_the_near_miss_floor(self) -> None:
        first, second = _job(), _job()
        first["id"], second["id"] = "job-1", "job-2"
        photo_duplicates = {
            "job-2": {"status": "no_match", "duplicate_of_job_id": "job-1",
                      "duplicate_of_filename": "first.pdf", "similarity": 0.11},
        }

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [first, second]}, photo_duplicates)))
        sheet = workbook["Duplicates"]

        self.assertEqual(sheet["B2"].value, None)  # no pair rows, notes only

    def test_duplicates_sheet_says_when_face_matching_never_ran(self) -> None:
        """An empty sheet reads as "nothing matched". When no identity model
        was available it means "nothing was compared", and only one of those
        is good news."""

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [_job()]})))
        notes = " ".join(
            str(row[0].value) for row in workbook["Duplicates"].iter_rows(min_row=2)
            if row[0].value
        )

        self.assertIn("Face matching did not run", notes)
        self.assertIn("No duplicate signal was found", notes)

    def test_summary_flags_a_document_on_both_sides_of_a_match(self) -> None:
        """Matching only ever looks backwards, so the earlier document of a
        matched pair would otherwise look untouched on the triage tab."""

        first, second = _job(), _job()
        first["id"], first["filename"] = "job-1", "first.pdf"
        second["id"], second["filename"] = "job-2", "second.pdf"
        photo_duplicates = {
            "job-2": {"status": "duplicate", "duplicate_of_job_id": "job-1",
                      "duplicate_of_filename": "first.pdf", "similarity": 0.87},
        }

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [first, second]}, photo_duplicates)))
        sheet = workbook["Summary"]
        headers = [cell.value for cell in sheet[1]]
        first_row = {header: cell.value for header, cell in zip(headers, sheet[2])}
        second_row = {header: cell.value for header, cell in zip(headers, sheet[3])}

        self.assertEqual(first_row["Duplicate signals"], "face")
        self.assertEqual(second_row["Duplicate signals"], "face")

    def test_per_page_sheet_breaks_each_check_down_by_page(self) -> None:
        """The point of the tab: page 1 carries the photo and the QR, page 2
        carries neither, and the sheet says so rather than reporting one count
        for the whole document."""

        job = _job()
        job["analyzers"].append("metadata")
        job["analyzer_runs"]["metadata"] = {"status": "completed", "error": None}
        job["results"]["metadata"] = normalize_analyzer_result("metadata", {
            "generation": {"kind": "scanner", "producer": "Scanner", "pages": 3},
            "document_metadata": {"created_at": "2026-01-02"},
        }).model_dump(mode="json")
        job["results"]["font_analysis"] = normalize_analyzer_result("font_analysis", {
            "unique_fonts": [
                {"xref": 4, "typeface": "Arial", "extension": "ttf",
                 "embedded": True, "pages_used": [1, 3]},
                {"xref": 5, "typeface": "Times", "extension": "ttf",
                 "embedded": True, "pages_used": [1]},
            ],
        }).model_dump(mode="json")

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Per Page"]
        headers = [cell.value for cell in sheet[1]]
        rows = [
            {header: cell.value for header, cell in zip(headers, row)}
            for row in sheet.iter_rows(min_row=2, max_row=4)
        ]

        self.assertEqual([row["Page"] for row in rows], [1, 2, 3])
        self.assertEqual([row["Embedded fonts"] for row in rows], [2, 0, 1])
        self.assertEqual(rows[0]["Font names"], "Arial, Times")
        self.assertEqual([row["Photos"] for row in rows], [1, 0, 0])
        self.assertEqual([row["Signatures"] for row in rows], [1, 0, 0])
        self.assertEqual([row["QR codes"] for row in rows], [1, 0, 0])

    def test_per_page_marks_a_check_that_never_ran_apart_from_a_zero(self) -> None:
        """0 is "looked, found nothing there". A check that was not run must not
        read the same way."""

        job = _job()
        job["analyzers"] = [a for a in job["analyzers"] if a != "signature"]
        del job["results"]["signature"]

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Per Page"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}

        self.assertEqual(row["Signatures"], "—")
        self.assertEqual(row["QR codes"], 1)

    def test_per_page_reports_whitener_probability_page_by_page(self) -> None:
        job = _job()
        job["analyzers"].append("tamper_scan")
        job["analyzer_runs"]["tamper_scan"] = {"status": "completed", "error": None}
        job["results"]["tamper_scan"] = normalize_analyzer_result("tamper_scan", {
            "document_probability": 0.62,
            "review_threshold": 0.35,
            "pages": [
                {"page": 1, "probability": 0.04, "regions": []},
                {"page": 2, "probability": 0.61, "regions": [{"kind": "whitener"}, {"kind": "whitener"}]},
            ],
        }).model_dump(mode="json")

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Per Page"]
        headers = [cell.value for cell in sheet[1]]
        cells = [
            {header: cell for header, cell in zip(headers, row)}
            for row in sheet.iter_rows(min_row=2, max_row=3)
        ]

        self.assertAlmostEqual(cells[0]["Whitener probability"].value, 0.04)
        self.assertAlmostEqual(cells[1]["Whitener probability"].value, 0.61)
        self.assertEqual(cells[1]["Whitener regions"].value, 2)
        # A percentage written as text would sort "9%" after "61%".
        self.assertEqual(cells[1]["Whitener probability"].number_format, "0.0%")

    def test_per_page_scores_a_page_by_its_worst_pairing(self) -> None:
        """Same-phone has no score for a page alone. The lowest pairing it took
        part in is how unlike the rest of the document it looked at its worst."""

        job = _job()
        job["results"]["same_phone"] = normalize_analyzer_result("same_phone", {
            "summary": {"overall_verdict": "possibly_different_phone_or_workflow",
                        "same_phone_score": 41},
            "pages": [{"page": 1}, {"page": 2}, {"page": 3}],
            "pair_assessments": [
                {"page_a": 1, "page_b": 2, "same_phone_score": 88},
                {"page_a": 1, "page_b": 3, "same_phone_score": 41},
                {"page_a": 2, "page_b": 3, "same_phone_score": 44},
            ],
        }).model_dump(mode="json")

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Per Page"]
        headers = [cell.value for cell in sheet[1]]
        scores = [
            dict(zip(headers, [cell.value for cell in row]))["Same-phone lowest pair score"]
            for row in sheet.iter_rows(min_row=2, max_row=4)
        ]

        self.assertEqual(scores, [41, 44, 41])

    def test_per_page_says_when_moire_carries_no_page_attribution(self) -> None:
        """The recapture scanner does not record which page a raster came from,
        and a blank column must not be read as a clean page."""

        job = _job()
        job["results"]["moire"] = normalize_analyzer_result("moire", {
            "file_verdict": "CLEAN",
            "images": [{"verdict": "CLEAN", "reason": "no anomalous periodic peaks"}],
        }).model_dump(mode="json")

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Per Page"]
        headers = [cell.value for cell in sheet[1]]
        row = {header: cell.value for header, cell in zip(headers, sheet[2])}
        notes = " ".join(
            str(cells[0].value) for cells in sheet.iter_rows(min_row=2) if cells[0].value
        )

        self.assertIn(row["Moire"], (None, ""))
        self.assertIn("does not record which page", notes)

    def test_a_document_that_did_not_run_a_check_is_left_off_that_sheet(self) -> None:
        job = _job()
        job["analyzers"] = [a for a in job["analyzers"] if a != "signature"]
        del job["results"]["signature"]

        workbook = load_workbook(BytesIO(batch_xlsx({"jobs": [job]})))
        sheet = workbook["Signature"]

        self.assertEqual(sheet.max_row, 1)  # header row only


if __name__ == "__main__":
    unittest.main()
